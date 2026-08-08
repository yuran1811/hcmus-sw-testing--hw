#!/usr/bin/env python3
"""Generate the HW06 reviewed case matrix and executable Postman assets."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


FIELDS = [
    "case_id", "api", "feature", "title", "origin", "audit_status",
    "audit_reason", "technique", "security_or_transition", "preconditions",
    "input_summary", "raw_body", "auth_mode", "prefail_count", "current_status",
    "expected_status", "check_type", "final_expected", "human_review",
]


def case(case_id, api, feature, title, *, origin="AI", audit="VALID",
         reason="The case is specific, executable, and traceable to the contract.",
         technique="Domain partition", coverage="None", pre="Fresh isolated SUT",
         body="{}", auth="none", prefail=0, current="", status=200,
         check="standard", expected="Response matches the specified status and schema."):
    display_origin = "Student-extension draft (HUMAN REQUIRED)" if origin == "Student" else origin
    return {
        "case_id": case_id,
        "api": api,
        "feature": feature,
        "title": title,
        "origin": display_origin,
        "audit_status": audit,
        "audit_reason": reason,
        "technique": technique,
        "security_or_transition": coverage,
        "preconditions": pre,
        "input_summary": body,
        "raw_body": body,
        "auth_mode": auth,
        "prefail_count": prefail,
        "current_status": current,
        "expected_status": status,
        "check_type": check,
        "final_expected": expected,
        "human_review": "TODO(HUMAN): confirm or correct this AI-assisted decision",
    }


def login_cases():
    api, feature = "POST /api/login", "FR-02"
    valid = lambda n: json.dumps({"email": f"login_{n:03d}@example.com", "password": "Valid123!"})
    rows = []
    def add(n, title, body=None, **kw):
        rows.append(case(f"LOGIN-{n:03d}", api, feature, title, body=body or valid(n), **kw))

    add(1, "Valid registered user", status=200, check="login_schema")
    add(2, "Email with leading and trailing spaces", json.dumps({"email":" login_002@example.com ","password":"Valid123!"}), status=400, expected="Reject unnormalized credentials with 400.")
    add(3, "Email case variation", json.dumps({"email":"LOGIN_003@EXAMPLE.COM","password":"Valid123!"}), status=401, expected="Return the documented generic authentication failure.")
    add(4, "Missing email", json.dumps({"password":"Valid123!"}), status=400)
    add(5, "Null email", json.dumps({"email":None,"password":"Valid123!"}), status=400)
    add(6, "Numeric email", json.dumps({"email":23127065,"password":"Valid123!"}), status=400)
    add(7, "Empty email", json.dumps({"email":"","password":"Valid123!"}), status=400)
    add(8, "Malformed email", json.dumps({"email":"not-an-email","password":"Valid123!"}), status=400)
    add(9, "Oversized email", json.dumps({"email":"a"*260+"@example.com","password":"Valid123!"}), status=400)
    add(10, "SQL injection in email", json.dumps({"email":"' OR 1=1 --","password":"Valid123!"}), status=401, coverage="SEC-01 injection")
    add(11, "Wrong password", json.dumps({"email":"login_011@example.com","password":"Wrong123!"}), status=401)
    add(12, "Empty password", json.dumps({"email":"login_012@example.com","password":""}), status=400)
    add(13, "Missing password", json.dumps({"email":"login_013@example.com"}), status=400)
    add(14, "Null password", json.dumps({"email":"login_014@example.com","password":None}), status=400)
    add(15, "Numeric password", json.dumps({"email":"login_015@example.com","password":12345678}), status=400)
    add(16, "Too-short password", json.dumps({"email":"login_016@example.com","password":"A1!"}), status=400)
    add(17, "Oversized password", json.dumps({"email":"login_017@example.com","password":"A1!"+"x"*1024}), status=400)
    add(18, "SQL injection in password", json.dumps({"email":"login_018@example.com","password":"' OR '1'='1"}), status=401, coverage="SEC-01 injection")
    add(19, "XSS payload in password", json.dumps({"email":"login_019@example.com","password":"<script>alert(1)</script>"}), status=401, coverage="SEC-02 output safety")
    add(20, "Empty JSON object", "{}", status=400)
    add(21, "Array instead of object", "[]", status=400, technique="Schema validation")
    add(22, "Unexpected extra field", json.dumps({"email":"login_022@example.com","password":"Valid123!","role":"admin"}), status=200, check="login_schema", coverage="SEC-04 mass assignment")
    add(23, "Unknown user", json.dumps({"email":"absent_023@example.com","password":"Valid123!"}), status=401)
    add(24, "Unicode local part", json.dumps({"email":"người@example.com","password":"Valid123!"}), status=401)
    add(25, "Password is case-sensitive", json.dumps({"email":"login_025@example.com","password":"valid123!"}), status=401)
    add(26, "One failed attempt", json.dumps({"email":"login_026@example.com","password":"Wrong123!"}), status=401, coverage="FR-02 lockout")
    add(27, "Correct password after two failures", valid(27), prefail=2, status=200, check="login_schema", coverage="FR-02 state transition", expected="Remain unlocked until three failed attempts, then authenticate.")
    add(28, "Correct password after three failures", valid(28), prefail=3, status=403, coverage="FR-02 state transition")
    add(29, "Wrong password after three failures", json.dumps({"email":"login_029@example.com","password":"Wrong123!"}), prefail=3, status=403, coverage="FR-02 state transition")
    add(30, "Immediate retry during lockout", valid(30), prefail=3, status=403, coverage="FR-02 state transition")
    add(31, "Successful response JSON schema", valid(31), status=200, check="login_schema", technique="Schema validation")
    add(32, "Successful response excludes plaintext password", valid(32), status=200, check="no_password", technique="Schema validation", coverage="SEC-07 sensitive data")
    add(33, "JSON content type on success", valid(33), status=200, check="json_content", technique="Schema validation")
    add(34, "Generic error prevents account enumeration", json.dumps({"email":"absent_034@example.com","password":"Wrong123!"}), status=401, check="generic_auth_error", coverage="SEC-06 enumeration")
    add(35, "Malformed candidate corrected to empty object", "{}", audit="INVALID", reason="The AI originally proposed syntactically broken JSON, which is a transport/parser test rather than a login-domain test; corrected to an empty valid object.", status=400)
    add(36, "Whitespace-only password", json.dumps({"email":"login_036@example.com","password":"   "}), origin="Student", status=400, coverage="SEC-05 input validation")
    add(37, "Duplicate JSON email key", '{"email":"attacker@example.com","email":"login_037@example.com","password":"Valid123!"}', origin="Student", audit="INCOMPLETE", reason="The AI omitted duplicate-key parsing; the final expectation documents last-key JSON parsing and successful authentication.", status=200, check="login_schema", technique="Parser ambiguity")
    add(38, "Prototype pollution-shaped field", json.dumps({"email":"login_038@example.com","password":"Valid123!","__proto__":{"role":"admin"}}), origin="Student", status=200, check="login_schema", coverage="SEC-04 object injection")
    add(39, "Very large harmless extra field", json.dumps({"email":"login_039@example.com","password":"Valid123!","padding":"x"*4096}), origin="Student", status=200, check="login_schema", technique="Robustness")
    add(40, "Student ID header retained on rejection", json.dumps({"email":"","password":""}), origin="Student", status=400, check="student_header", coverage="Anti-cheat evidence")
    return rows


def checkout_cases():
    api, feature = "POST /api/checkout", "FR-08"
    rows = []
    def add(n, title, body, **kw):
        rows.append(case(f"CHECKOUT-{n:03d}", api, feature, title, body=body, auth=kw.pop("auth", "user"), **kw))
    j = json.dumps
    add(1,"Valid checkout",j({"total_amount":200000,"shipping_address":"123 Le Loi, TP.HCM"}),status=200,check="checkout_schema")
    add(2,"Minimum positive total",j({"total_amount":1,"shipping_address":"A"}),status=200,check="checkout_schema")
    add(3,"Large valid integer total",j({"total_amount":2147483647,"shipping_address":"A"}),status=200,check="checkout_schema")
    add(4,"Zero total",j({"total_amount":0,"shipping_address":"A"}),status=400)
    add(5,"Negative total",j({"total_amount":-1,"shipping_address":"A"}),status=400)
    add(6,"Decimal total",j({"total_amount":10.5,"shipping_address":"A"}),status=400)
    add(7,"Numeric string total",j({"total_amount":"200000","shipping_address":"A"}),status=400)
    add(8,"Nonnumeric total",j({"total_amount":"NaN","shipping_address":"A"}),status=400)
    add(9,"Null total",j({"total_amount":None,"shipping_address":"A"}),status=400)
    add(10,"Missing total",j({"shipping_address":"A"}),status=400)
    add(11,"Boolean total",j({"total_amount":True,"shipping_address":"A"}),status=400)
    add(12,"Array total",j({"total_amount":[200000],"shipping_address":"A"}),status=400)
    add(13,"Object total",j({"total_amount":{"value":200000},"shipping_address":"A"}),status=400)
    add(14,"Integer above safe range",j({"total_amount":9007199254740992,"shipping_address":"A"}),status=400)
    add(15,"SQL injection-shaped total",j({"total_amount":"0); DROP TABLE orders;--","shipping_address":"A"}),status=400,coverage="SEC-01 injection")
    add(16,"Vietnamese shipping address",j({"total_amount":200000,"shipping_address":"12 Nguyễn Huệ, Quận 1"}),status=200,check="checkout_schema")
    add(17,"Empty shipping address",j({"total_amount":200000,"shipping_address":""}),status=400)
    add(18,"Missing shipping address",j({"total_amount":200000}),status=400)
    add(19,"Null shipping address",j({"total_amount":200000,"shipping_address":None}),status=400)
    add(20,"Whitespace shipping address",j({"total_amount":200000,"shipping_address":"   "}),status=400)
    add(21,"Numeric shipping address",j({"total_amount":200000,"shipping_address":123}),status=400)
    add(22,"Oversized shipping address",j({"total_amount":200000,"shipping_address":"x"*2048}),status=400)
    add(23,"XSS shipping address",j({"total_amount":200000,"shipping_address":"<script>alert(1)</script>"}),status=400,coverage="SEC-02 stored XSS")
    add(24,"SQL injection shipping address",j({"total_amount":200000,"shipping_address":"'; DROP TABLE orders;--"}),status=400,coverage="SEC-01 injection")
    add(25,"Unexpected role field",j({"total_amount":200000,"shipping_address":"A","role":"admin"}),status=200,check="checkout_schema",coverage="SEC-04 mass assignment")
    add(26,"Missing bearer token",j({"total_amount":200000,"shipping_address":"A"}),auth="missing",status=401,coverage="SEC-03 authentication")
    add(27,"Malformed bearer token",j({"total_amount":200000,"shipping_address":"A"}),auth="invalid",status=403,coverage="SEC-03 authentication")
    add(28,"Admin token checkout",j({"total_amount":200000,"shipping_address":"A"}),auth="admin",status=200,check="checkout_schema")
    add(29,"Tampered bearer token",j({"total_amount":200000,"shipping_address":"A"}),auth="tampered",status=403,coverage="SEC-03 authentication")
    add(30,"Empty JSON object",j({}),status=400)
    add(31,"Array instead of object",j([]),status=400,technique="Schema validation")
    add(32,"Successful response schema",j({"total_amount":200000,"shipping_address":"A"}),status=200,check="checkout_schema",technique="Schema validation")
    add(33,"JSON response content type",j({"total_amount":200000,"shipping_address":"A"}),status=200,check="json_content",technique="Schema validation")
    add(34,"AI case corrected from client-supplied user_id",j({"total_amount":200000,"shipping_address":"A","user_id":999999}),audit="INVALID",reason="The AI expected the server to trust user_id; corrected to require identity from the bearer token and ignore the extra field.",status=200,check="checkout_schema",coverage="SEC-04 identity binding")
    add(35,"Incomplete idempotency case corrected",j({"total_amount":200000,"shipping_address":"A"}),audit="INCOMPLETE",reason="The specification defines no idempotency key; corrected to a single successful creation with a unique orderId.",status=200,check="checkout_schema")
    add(36,"Checkout with empty cart",j({"total_amount":200000,"shipping_address":"A"}),origin="Student",status=400,check="no_empty_cart",coverage="FR-08 business rule")
    add(37,"Client total does not match cart",j({"total_amount":1,"shipping_address":"A"}),origin="Student",status=400,coverage="FR-08 price integrity")
    add(38,"Negative total bypass attempt",j({"total_amount":-999999999,"shipping_address":"A"}),origin="Student",status=400,coverage="SEC-05 integrity")
    add(39,"Foreign user_id is ignored",j({"total_amount":200000,"shipping_address":"A","user_id":1}),origin="Student",status=200,check="checkout_schema",coverage="SEC-04 IDOR")
    add(40,"Exponent overflow total",'{"total_amount":1e309,"shipping_address":"A"}',origin="Student",status=400,technique="Numeric parser boundary")
    return rows


def order_cases():
    api, feature = "PUT /api/admin/orders/:id/status", "FR-18"
    rows = []
    def add(n, title, current, target, **kw):
        body = "{}" if target == "__missing__" else json.dumps({"status": target})
        rows.append(case(f"ORDER-{n:03d}", api, feature, title, body=body,
                         auth=kw.pop("auth", "admin"), current=current,
                         technique=kw.pop("technique", "State transition"),
                         coverage=kw.pop("coverage", f"{current} -> {target}"), **kw))
    add(1,"Pending to confirmed","pending","confirmed",status=200,check="order_schema")
    add(2,"Pending to canceled","pending","canceled",status=200,check="order_schema")
    add(3,"Pending to shipping","pending","shipping",status=400)
    add(4,"Pending to delivered","pending","delivered",status=400)
    add(5,"Pending to pending","pending","pending",status=400)
    add(6,"Confirmed to shipping","confirmed","shipping",status=200,check="order_schema")
    add(7,"Confirmed to canceled","confirmed","canceled",status=200,check="order_schema")
    add(8,"Confirmed to pending","confirmed","pending",status=400)
    add(9,"Confirmed to delivered","confirmed","delivered",status=400)
    add(10,"Confirmed to confirmed","confirmed","confirmed",status=400)
    add(11,"Shipping to delivered","shipping","delivered",status=200,check="order_schema")
    add(12,"Shipping to canceled","shipping","canceled",status=400)
    add(13,"Shipping to confirmed","shipping","confirmed",status=400)
    add(14,"Shipping to pending","shipping","pending",status=400)
    add(15,"Shipping to shipping","shipping","shipping",status=400)
    add(16,"Delivered to pending","delivered","pending",status=400)
    add(17,"Delivered to confirmed","delivered","confirmed",status=400)
    add(18,"Delivered to shipping","delivered","shipping",status=400)
    add(19,"Delivered to canceled","delivered","canceled",status=400)
    add(20,"Delivered to delivered","delivered","delivered",status=400)
    add(21,"Canceled to delivered","canceled","delivered",status=400,coverage="FR-10 terminal-state integrity")
    add(22,"Canceled to pending","canceled","pending",status=400)
    add(23,"Canceled to confirmed","canceled","confirmed",status=400)
    add(24,"Canceled to shipping","canceled","shipping",status=400)
    add(25,"Canceled to canceled","canceled","canceled",status=400)
    add(26,"Unknown status enum","pending","refunded",status=400)
    add(27,"Empty status","pending","",status=400)
    add(28,"Null status","pending",None,status=400)
    add(29,"Missing status","pending","__missing__",status=400)
    add(30,"Numeric status","pending",23127065,status=400)
    add(31,"Array status","pending",["confirmed"],status=400)
    add(32,"Object status","pending",{"value":"confirmed"},status=400)
    add(33,"Nonexistent order","nonexistent","confirmed",status=404)
    add(34,"Missing bearer token","pending","confirmed",auth="missing",status=401,coverage="SEC-03 authentication")
    add(35,"Malformed bearer token","pending","confirmed",auth="invalid",status=403,coverage="SEC-03 authentication")
    add(36,"Normal user changes admin-managed order","pending","confirmed",origin="Student",auth="user",status=403,coverage="SEC-04 role escalation")
    add(37,"Normal user changes another user's order","pending","canceled",origin="Student",auth="user",status=403,coverage="SEC-04 IDOR")
    add(38,"Successful response schema","pending","confirmed",origin="Student",status=200,check="order_schema",technique="Schema validation")
    add(39,"SQL injection status","pending","confirmed' OR '1'='1",origin="Student",status=400,coverage="SEC-01 injection")
    add(40,"Normal user attempts mutation of a delivered order","delivered","canceled",origin="Student",auth="user",audit="INCOMPLETE",reason="The AI considered only the invalid transition and omitted the role boundary; corrected so authorization is evaluated first and a normal user receives 403.",status=403,coverage="SEC-04 role check before FR-10 transition")
    return rows


COLLECTION_PRE = [
    'pm.request.headers.upsert({key: "X-Student-Id", value: "23127065"});',
    'console.log("X-Student-Id: 23127065");',
]


LOGIN_PRE = r'''const base = pm.environment.get("baseUrl");
const id = pm.iterationData.get("case_id");
const email = `login_${String(id).split("-")[1]}@example.com`;
pm.request.body.raw = pm.iterationData.get("raw_body");
pm.request.headers.upsert({key:"Content-Type", value:"application/json"});
const registration = {url:`${base}/api/register`, method:"POST", header:{"Content-Type":"application/json","X-Student-Id":"23127065"}, body:{mode:"raw",raw:JSON.stringify({name:id,email,password:"Valid123!"})}};
pm.sendRequest(registration, function () {
  let remaining = Number(pm.iterationData.get("prefail_count") || 0);
  const failOnce = () => {
    if (remaining-- <= 0) return;
    pm.sendRequest({url:`${base}/api/login`,method:"POST",header:{"Content-Type":"application/json","X-Student-Id":"23127065"},body:{mode:"raw",raw:JSON.stringify({email,password:"Wrong123!"})}}, failOnce);
  };
  failOnce();
});'''


AUTH_PRE = r'''const base = pm.environment.get("baseUrl");
pm.request.body.raw = pm.iterationData.get("raw_body");
pm.request.headers.upsert({key:"Content-Type", value:"application/json"});
const mode = pm.iterationData.get("auth_mode");
if (mode === "missing") { pm.request.headers.remove("Authorization"); }
else if (mode === "invalid") { pm.request.headers.upsert({key:"Authorization",value:"Bearer not-a-jwt"}); }
else {
  const credentials = mode === "admin" ? {email:"admin@eshop.com",password:"Admin123!"} : {email:"test@eshop.com",password:"Test1234!"};
  pm.sendRequest({url:`${base}/api/login`,method:"POST",header:{"Content-Type":"application/json","X-Student-Id":"23127065"},body:{mode:"raw",raw:JSON.stringify(credentials)}}, (err,res) => {
    let token = res && res.json().token;
    if (mode === "tampered") token = `${token}x`;
    pm.request.headers.upsert({key:"Authorization",value:`Bearer ${token}`});
  });
}'''


ORDER_PRE = r'''const base = pm.environment.get("baseUrl");
pm.request.body.raw = pm.iterationData.get("raw_body");
pm.request.headers.upsert({key:"Content-Type", value:"application/json"});
const login = (email,password,next) => pm.sendRequest({url:`${base}/api/login`,method:"POST",header:{"Content-Type":"application/json","X-Student-Id":"23127065"},body:{mode:"raw",raw:JSON.stringify({email,password})}},(e,r)=>next(r.json().token));
login("test@eshop.com","Test1234!", userToken => login("admin@eshop.com","Admin123!", adminToken => {
  const mode = pm.iterationData.get("auth_mode");
  if (mode === "missing") pm.request.headers.remove("Authorization");
  else if (mode === "invalid") pm.request.headers.upsert({key:"Authorization",value:"Bearer not-a-jwt"});
  else pm.request.headers.upsert({key:"Authorization",value:`Bearer ${mode === "user" ? userToken : adminToken}`});
  if (pm.iterationData.get("current_status") === "nonexistent") { pm.variables.set("order_id","99999999"); return; }
  pm.sendRequest({url:`${base}/api/checkout`,method:"POST",header:{"Content-Type":"application/json","Authorization":`Bearer ${userToken}`,"X-Student-Id":"23127065"},body:{mode:"raw",raw:JSON.stringify({total_amount:200000,shipping_address:"Setup address"})}},(e,r)=>{
    const orderId = r.json().orderId; pm.variables.set("order_id",String(orderId));
    const paths = {pending:[],confirmed:["confirmed"],shipping:["confirmed","shipping"],delivered:["confirmed","shipping","delivered"],canceled:["canceled"]};
    const states = paths[pm.iterationData.get("current_status")] || [];
    const advance = () => { const state=states.shift(); if (!state) return; pm.sendRequest({url:`${base}/api/admin/orders/${orderId}/status`,method:"PUT",header:{"Content-Type":"application/json","Authorization":`Bearer ${adminToken}`,"X-Student-Id":"23127065"},body:{mode:"raw",raw:JSON.stringify({status:state})}},advance); }; advance();
  });
}));'''


TEST_SCRIPT = r'''const id = pm.iterationData.get("case_id");
const expected = Number(pm.iterationData.get("expected_status"));
pm.test(`${id} expected HTTP ${expected}`, () => pm.response.to.have.status(expected));
pm.test(`${id} sent student header`, () => pm.expect(pm.request.headers.get("X-Student-Id")).to.eql("23127065"));
const check = pm.iterationData.get("check_type");
let body = {}; try { body = pm.response.json(); } catch (_) {}
if (check === "login_schema" && pm.response.code === 200) pm.test(`${id} login schema`, () => { pm.expect(body).to.have.keys("message","token","user"); pm.expect(body.token).to.be.a("string"); });
if (check === "checkout_schema" && pm.response.code === 200) pm.test(`${id} checkout schema`, () => { pm.expect(body).to.have.keys("message","orderId"); pm.expect(body.orderId).to.be.a("number"); });
if (check === "order_schema" && pm.response.code === 200) pm.test(`${id} order schema`, () => pm.expect(body).to.eql({message:"Order status updated"}));
if (check === "no_password" && pm.response.code === 200) pm.test(`${id} excludes password`, () => pm.expect(body.user).to.not.have.property("password"));
if (check === "json_content") pm.test(`${id} JSON content type`, () => pm.expect(pm.response.headers.get("Content-Type")).to.include("application/json"));
if (check === "generic_auth_error") pm.test(`${id} generic auth error`, () => pm.expect(body.error).to.eql("Invalid email or password"));
if (check === "student_header") pm.test(`${id} exact anti-cheat header`, () => pm.expect(pm.request.headers.get("X-Student-Id")).to.eql("23127065"));'''


def event(kind, script):
    lines = script if isinstance(script, list) else script.splitlines()
    return {"listen": kind, "script": {"type": "text/javascript", "exec": lines}}


def request_item(name, method, url, pre):
    return {
        "name": name,
        "event": [event("prerequest", pre), event("test", TEST_SCRIPT)],
        "request": {
            "method": method,
            "header": [{"key":"Content-Type","value":"application/json"}],
            "body": {"mode":"raw","raw":"{{raw_body}}","options":{"raw":{"language":"json"}}},
            "url": url,
            "description": "Data-driven execution; case metadata comes from the matching external JSON file.",
        },
    }


def write_collection(out, student_id):
    collection = {
        "info": {"name": f"{student_id} HW06 API Testing", "schema":"https://schema.getpostman.com/json/collection/v2.1.0/collection.json", "description":"Pool A login, Pool B checkout, and Pool C admin order-state testing."},
        "event": [event("prerequest", [x.replace("23127065", student_id) for x in COLLECTION_PRE])],
        "variable": [{"key":"studentId","value":student_id,"type":"string"}],
        "item": [
            {"name":"Pool A - Login","item":[request_item("Execute Login Case","POST",{"raw":"{{baseUrl}}/api/login","host":["{{baseUrl}}"],"path":["api","login"]},LOGIN_PRE.replace("23127065", student_id))]},
            {"name":"Pool B - Checkout","item":[request_item("Execute Checkout Case","POST",{"raw":"{{baseUrl}}/api/checkout","host":["{{baseUrl}}"],"path":["api","checkout"]},AUTH_PRE.replace("23127065", student_id))]},
            {"name":"Pool C - Admin Order Status","item":[request_item("Execute Order Status Case","PUT",{"raw":"{{baseUrl}}/api/admin/orders/{{order_id}}/status","host":["{{baseUrl}}"],"path":["api","admin","orders","{{order_id}}","status"]},ORDER_PRE.replace("23127065", student_id))]},
        ],
    }
    (out / "postman").mkdir(parents=True, exist_ok=True)
    (out / "postman" / f"{student_id}_HW06.postman_collection.json").write_text(json.dumps(collection, ensure_ascii=False, indent=2)+"\n")
    environment = {"id":f"{student_id}-localhost","name":"HW06 Localhost","values":[{"key":"baseUrl","value":"http://127.0.0.1:3000","enabled":True,"type":"default"}],"_postman_variable_scope":"environment","_postman_exported_using":"Codex + Postman Collection v2.1"}
    (out / "postman" / "localhost.postman_environment.json").write_text(json.dumps(environment, ensure_ascii=False, indent=2)+"\n")


def write_workbook(out, groups):
    wb = Workbook(); wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    for name, rows in groups.items():
        ws = wb.create_sheet(name)
        ws.append(FIELDS)
        for cell in ws[1]: cell.font=Font(color="FFFFFF",bold=True); cell.fill=fill
        for row in rows: ws.append([row.get(k,"") for k in FIELDS])
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(55,max(12,max(len(str(c.value or "")) for c in col)+2))
            for c in col: c.alignment=Alignment(vertical="top",wrap_text=True)
    ws=wb.create_sheet("Summary"); ws.append(["API","AI","Student","Total","VALID","INVALID","INCOMPLETE"])
    for name,rows in groups.items():
        origins=Counter("Student" if r["origin"].startswith("Student") else r["origin"] for r in rows); audits=Counter(r["audit_status"] for r in rows)
        ws.append([name,origins["AI"],origins["Student"],len(rows),audits["VALID"],audits["INVALID"],audits["INCOMPLETE"]])
    ws.append(["TOTAL",105,15,120,"","",""])
    (out/"test-cases").mkdir(parents=True,exist_ok=True)
    wb.save(out/"test-cases"/"23127065_HW06_Test_Cases.xlsx")


def write_manifest(out, groups):
    lines=["# Generated test-case manifest","", "| API | AI-generated | Student-extension drafts | Total |", "|---|---:|---:|---:|"]
    for name,rows in groups.items():
        origins=Counter("Student" if r["origin"].startswith("Student") else r["origin"] for r in rows); lines.append(f"| {name} | {origins['AI']} | {origins['Student']} | {len(rows)} |")
    lines += ["| **Total** | **105** | **15** | **120** |","", "Every row in the Excel workbook includes the original audit label, reasoning, corrected final expectation, and a `TODO(HUMAN)` review field. Machine execution does not convert those human-review markers into claims of student authorship."]
    (out/"test-cases"/"Test_Summary.md").write_text("\n".join(lines)+"\n")


def main():
    parser=argparse.ArgumentParser(); parser.add_argument("--student-id",required=True); parser.add_argument("--output",type=Path,required=True); args=parser.parse_args()
    out=args.output.resolve(); groups={"Login":login_cases(),"Checkout":checkout_cases(),"OrderStatus":order_cases()}
    for name,rows in groups.items():
        assert len(rows)==40 and sum(r["origin"]=="AI" for r in rows)==35 and sum(r["origin"].startswith("Student") for r in rows)==5
    (out/"test-data").mkdir(parents=True,exist_ok=True)
    for name,rows in groups.items():
        (out/"test-data"/f"{name.lower()}-cases.json").write_text(json.dumps(rows,ensure_ascii=False,indent=2)+"\n")
    (out/"ci-data").mkdir(parents=True,exist_ok=True)
    for name,rows in groups.items():
        (out/"ci-data"/f"{name.lower()}-pass.json").write_text(json.dumps([rows[0]],ensure_ascii=False,indent=2)+"\n")
    intentional_failure = dict(groups["Login"][0])
    intentional_failure.update({"case_id":"CI-DEMO-FAIL-001","title":"Intentional CI failure demonstration","expected_status":418,"final_expected":"Deliberately impossible expectation used only to demonstrate a red pipeline."})
    (out/"ci-data"/"login-intentional-fail.json").write_text(json.dumps([intentional_failure],ensure_ascii=False,indent=2)+"\n")
    write_workbook(out,groups); write_manifest(out,groups); write_collection(out,args.student_id)
    print(json.dumps({"student_id":args.student_id,"apis":3,"ai_generated":105,"student_extension_drafts":15,"total":120,"output":str(out)},indent=2))


if __name__ == "__main__": main()
