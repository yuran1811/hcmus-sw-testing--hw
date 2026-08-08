#!/usr/bin/env python3
"""Fail-closed structural audit for the 23127065 HW06 submission."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REQUIRED = [
    "README.md", "Main_Report.md", "Main_Report.pdf", "AI_Audit_Report.md",
    "AI_Audit_Report.pdf", "AI_Critique.md", "AI_Critique.pdf", "Bug_Report.md",
    "HUMAN_ACTION_REQUIRED.md", "Git_Commit_Log.txt", "test-cases/23127065_HW06_Test_Cases.xlsx",
    "postman/23127065_HW06.postman_collection.json",
    "postman/localhost.postman_environment.json", "package.json", "package-lock.json",
    "newman/results/login.json", "newman/results/checkout.json",
    "newman/results/orderstatus.json", "newman/reports/login/index.html",
    "newman/reports/checkout/index.html", "newman/reports/orderstatus/index.html",
    "ci/CI_CD_Report.md", "agent-skill/generate-api-tests/SKILL.md",
    "agent-skill/generate-api-tests/design/pseudocode.md",
]


def fail(message):
    print(f"FAIL: {message}")
    return 1


def main():
    errors = 0
    for rel in REQUIRED:
        path = ROOT / rel
        if not path.is_file() or path.stat().st_size == 0:
            errors += fail(f"missing or empty {rel}")

    groups = {"login": "LOGIN", "checkout": "CHECKOUT", "orderstatus": "ORDER"}
    for filename, prefix in groups.items():
        rows = json.loads((ROOT / "test-data" / f"{filename}-cases.json").read_text())
        ai = sum(r["origin"] == "AI" for r in rows)
        drafts = sum(r["origin"].startswith("Student-extension draft") for r in rows)
        if len(rows) != 40 or ai != 35 or drafts != 5:
            errors += fail(f"{filename}: expected 40/35/5, got {len(rows)}/{ai}/{drafts}")
        if len({r["case_id"] for r in rows}) != 40 or any(not r["case_id"].startswith(prefix) for r in rows):
            errors += fail(f"{filename}: IDs are not 40 unique {prefix} IDs")

    collection_text = (ROOT / "postman" / "23127065_HW06.postman_collection.json").read_text()
    if "X-Student-Id" not in collection_text or "23127065" not in collection_text:
        errors += fail("collection lacks mandatory student header")
    json.loads(collection_text)

    expected_stats = {"login": (40, 90, 20), "checkout": (40, 91, 26), "orderstatus": (40, 86, 4)}
    for name, expected in expected_stats.items():
        run = json.loads((ROOT / "newman" / "results" / f"{name}.json").read_text())["run"]
        actual = (run["stats"]["iterations"]["total"], run["stats"]["assertions"]["total"], run["stats"]["assertions"]["failed"])
        if actual != expected:
            errors += fail(f"{name}: execution stats {actual} != {expected}")

    wb = load_workbook(ROOT / "test-cases" / "23127065_HW06_Test_Cases.xlsx", read_only=True)
    for sheet in ("Login", "Checkout", "OrderStatus"):
        if sheet not in wb.sheetnames or wb[sheet].max_row != 41:
            errors += fail(f"workbook sheet {sheet} does not contain header + 40 cases")

    human_markers = []
    for path in ROOT.rglob("*.md"):
        count = path.read_text(errors="replace").count("TODO(HUMAN)")
        if count:
            human_markers.append((str(path.relative_to(ROOT)), count))

    print("PASS: machine-verifiable structure and recorded Newman totals" if not errors else f"FAILURES: {errors}")
    print("HUMAN-REQUIRED markers (expected until student completion):")
    for path, count in sorted(human_markers):
        print(f"  {path}: {count}")
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
